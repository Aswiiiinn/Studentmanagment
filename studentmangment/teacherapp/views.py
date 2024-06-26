
from datetime import date
import random
import string
from django.http import HttpResponse
from django.shortcuts import render,redirect
import openpyxl
from studentapp.models import Teacher
from adminapp.forms import TeacherForm
from django.core.mail import send_mail
from django.views import View
from .models import *
from studentapp.models import *
from adminapp.forms import *
from . forms import *

from studentmangment import settings

class choose(View):
    def get(self,request):
        return render(request,'teacherdash.html')
    def post(self,request):
        name  = request.POST.get('action')
        if name == "Add_Course_Material":
            return redirect('teacherapp:addMeterial')
        if name == "Add_Mark":
            return redirect('teacherapp:Scoredetailsview')
        if name == "Add_Attendence":
            return redirect("")
        if name == "Add_Excel":
            return redirect('teacherapp:exceldownload')

class viewteacher(View):
    def get(self,request):
        item = Teacher.objects.all()
        form = TeacherForm()
        
        context = {'form': form, 'item': item}
        return render(request,'teacherview.html',context)
class addteacher(View):
    def get(self,request):
        form = TeacherForm()
        return render(request,'addteacher.html',{'form':form})
    def post(self,request):
        form = TeacherForm(request.POST)
        
        if form.is_valid():
            teacher =form.save(commit=False)
            passwordlength = 6
            random_password =''.join(random.choices(string.ascii_letters+string.digits,k=passwordlength))
            
            teacher.username = teacher.name
            teacher.password =random_password
            email = teacher.email
            
            subject  = 'your account Info'
            message =f'hello this is your username\n: {teacher.username}\n this is your password\n:{teacher.password}'
            from_email =settings.EMAIL_HOST_USER
            to_email = [email]
            send_mail(subject,message,from_email,to_email)
            teacher.save()
            
            
            return redirect('teacherapp:teacher_view')
        return render(request,'addteacher.html',{'form':form})
class updateteacher(View):
    def get(self,request,pk):
        item = Teacher.objects.get(id=pk)
        form = TeacherForm(instance=item)
        # item2 = Teacher.objects.filter(batches__name='A').annotate(scores = Max('teacher__student_score'))
        # for teacher in item2:
        #     print(teacher.name)
        #     for batch in teacher.batches.all():
        #      print(batch.start_date)
        #      print(batch.end_date)

    
        return render(request,'updateteacher.html',{'form':form})
    
    def post(self,request,pk):
        item = Teacher.objects.get(id=pk)
        form = TeacherForm(request.POST,instance=item)
        if form.is_valid():
            form.save()
            return redirect('teacherapp:teacher_view')
        return render(request,'updateteacher.html',{'form':form})
class delete(View):
    def get(self,request,pk):
        item = Teacher.objects.get(id=pk)
        item.delete()
        return redirect('teacherapp:teacher_view')
class addMeterial(View):
    def get (self,request):
        form = courseForm()
        return render(request,'addmeterial.html',{'form':form})
    
    def post(self,request):
        form = courseForm(request.POST,request.FILES)

        if form.is_valid():
            meterial_instance = form.save(commit=False)
            meterial_instance.save()
            return redirect('teacherapp:meterial_view')
        return render(request,'addmeterial.html',{'form':form})
class tchrdash(View):
    def get(self,request):
        return render(request,'teacherdash.html')
    
class meterialview(View):
    def get(self,request):
        item  = CourseMeterial.objects.all()
        form = courseForm()
        context = {'form': form, 'item': item}
        return render(request,'meterialview.html',context)
class deleteview(View):
    def get(self,request,pk):
        item = CourseMeterial.objects.get(id=pk)
        item.delete()
        return redirect('teacherapp:meterial_view')
class exceldownload(View):
    def get(self,request):
        form = importForm()
        return render(request,'exceldownload.html',{'form': form})
    def post(self,request):
        form = importForm(request.POST,request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            excel = wb.active
            for rows in excel.iter_rows(min_row=3,values_only=True):
                importex.objects.create(
                    Student_Name =rows[0],
                    Exam_Name =rows[1],
                    Date =rows[2],
                    Point=rows[3] )
                
            return HttpResponse('succsess')
        return render(request,'exceldownload.html',{'form': form})
class ExportToExcelView(View):
    def get(self, request, *args, **kwargs):
        # Create a workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Excel Upload Data"

        # Define the headers
        headers = ['Student_Name', 'Exam_Name', 'Date', 'Point']
        ws.append(headers)

        # Fetch the data from the database
        excel_uploads = importex.objects.all()

        # Append the data to the worksheet
        for upload in excel_uploads:
            ws.append([upload.Student_Name, upload.Exam_Name, upload.Date, upload.Point])

        # Create an HTTP response with the appropriate headers for Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=excelupload_data.xlsx'

        # Save the workbook to the response
        wb.save(response)

        # Return the response
        return response
class Scoredetailsview(View):
    def get(self, request):
        item = Achivments.objects.all()
        return render(request,'achivements.html',{'item':item})

class Scoredetailsviewadd(View):
    def get(self,request):
        form = achivmentForm()
        return render(request,'addchivements.html',{'form':form})
    def post(self,request):
        form = achivmentForm(request.POST)
        if form.is_valid():
            achivment_instance = form.save(commit=False)
            score = int(request.POST.get('score'))
            if score >= 90:
                achivment_instance.Grade = 'S'
                achivment_instance.class_perfomance ="outstanding"
            elif score >=80 and score<90:
                achivment_instance.Grade = 'A+'
                achivment_instance.class_perfomance ='out'
            elif score >=70 and score<80:
                achivment_instance.Grade = 'A'
                achivment_instance.class_perfomance ='veryGood'
            elif score >=60 and score<70:
                achivment_instance.Grade = 'B+'
                achivment_instance.class_perfomance ='good'

            elif score >=50 and score<60:
                achivment_instance.Grade = 'B'
                achivment_instance.class_perfomance ='average'
                
            elif score >=40 and score<50:
                achivment_instance.Grade = 'C+'
                achivment_instance.class_perfomance ='bad'
                
            elif score >=30 and score<40:
                achivment_instance.Grade = 'C'
                achivment_instance.class_perfomance ='bad'
            else:
                achivment_instance.Grade = 'F'
                achivment_instance.class_perfomance ='veryBad'
        
            achivment_instance.save()
        return redirect('teacherapp:Scoredetailsview')        
class updatemark(View):
    def get(self, request,pk):
        item = Achivments.objects.get(id=pk)
        form = achivmentForm(instance=item)
        return render(request,'updatemark.html',{'form':form})
    def post(self, request,pk):
        item = Achivments.objects.get(id=pk)
        form = achivmentForm(request.POST,instance=item)
        if form.is_valid():
            achivment_instance = form.save(commit=False)
            score = int(request.POST.get('score'))
            if score >= 90:
                achivment_instance.Grade = 'S'
                achivment_instance.class_perfomance ="outstanding"
            elif score >=80 and score<90:
                achivment_instance.Grade = 'A+'
                achivment_instance.class_perfomance ='out'
            elif score >=70 and score<80:
                achivment_instance.Grade = 'A'
                achivment_instance.class_perfomance ='veryGood'
            elif score >=60 and score<70:
                achivment_instance.Grade = 'B+'
                achivment_instance.class_perfomance ='good'

            elif score >=50 and score<60:
                achivment_instance.Grade = 'B'
                achivment_instance.class_perfomance ='average'
                
            elif score >=40 and score<50:
                achivment_instance.Grade = 'C+'
                achivment_instance.class_perfomance ='bad'
                
            elif score >=30 and score<40:
                achivment_instance.Grade = 'C'
                achivment_instance.class_perfomance ='bad'
            else:
                achivment_instance.Grade = 'F'
                achivment_instance.class_perfomance ='veryBad'
        
            achivment_instance.save()
            
            return redirect('teacherapp:Scoredetailsview')
        return render(request,'updatemark.html',{'form':form})
class deleteviewscore(View):
    def get(self,request,pk):
        item = Achivments.objects.get(id=pk)
        item.delete()
        return redirect('teacherapp:Scoredetailsview')
class attandenceview(View):
    def get(self,request):
        form =attandenceForm()
        return render(request,'updateattandence.html',{'form':form})
    def post(self,request):
        form = attandenceForm(request.POST)
        if form.is_valid():
            teacher_id = form.cleaned_data['teacher_id']
            present = form.cleaned_data['present']
            arrival_time = form.cleaned_data['arrival_time']
            break_start = form.cleaned_data['break_start']
            break_end= form.cleaned_data['break_end']
            departure_time = form.cleaned_data['departure_time']
             
            attanjdence_date = date.today()
            attandence1.objects.create(teacher_id=teacher_id,present=present,date=attanjdence_date,arrival_time =arrival_time, break_start=break_start,break_end=break_end, departure_time= departure_time)
            return redirect('teacherapp:attandencelistview')
        return render(request,'updateattandence.html',{'form':form})
class attandencelistview(View):
    def get(self,request):
        item = attandence1.objects.all()
        return render(request,'attandenceview.html',{'item':item})
class updateattandenceview(View):
    def get (self,request,pk):
        item = attandence1.objects.get(id=pk)
        form =attandenceForm(instance = item)
        return render(request,'updateattandence.html',{'form':form})
    def post(self,request,pk):
        item = attandence1.objects.get(id=pk)
        form = attandenceForm(request.POST,instance=item)
        if form.is_valid():
            form.save()
            return redirect('teacherapp:attandencelistview')
        return render(request,'updateattandence.html',{'form':form})
class delete_attandenceview(View):
    def get(self,request,pk):
        item = attandence1.objects.get(id=pk)
        item.delete()
        return redirect('teacherapp:attandenceview')
        